from pathlib import Path
import pandas as pd
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl import load_workbook

class ReportGenerator:
    def __init__(self, processor, output_dir="reports"):
        self.processor = processor
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)

    def generate_csv_report(self):
        summary = self.processor.summary()
        df = pd.DataFrame([summary])
        file_path = self.output_dir / "financial_summary.csv"
        df.to_csv(file_path, index=False)
        return file_path

    def generate_excel_report(self):
        file_path = self.output_dir / "financial_report.xlsx"

        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            # --- SUMMARY (VERTICAL) ---
            summary_dict = self.processor.summary()

            summary_df = pd.DataFrame(
                list(summary_dict.items()),
                columns=["Metric", "Value"]
            )

            summary_df.to_excel(writer, sheet_name="Summary", index=False)

            # --- EXPENSES BY CATEGORY ---
            expenses_df = self.processor.expenses_by_category().reset_index()
            expenses_df.to_excel(writer, sheet_name="Expenses by Category", index=False)


        workbook = load_workbook(file_path)

        # --- PIE CHART ---
        sheet = workbook["Expenses by Category"]

        pie = PieChart()
        pie.title = "Expenses by Category"

        data = Reference(sheet, min_col=2, min_row=1, max_row=sheet.max_row)
        labels = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)

        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)

        sheet.add_chart(pie, "E2")

        # --- BAR CHART ---
        summary_sheet = workbook["Summary"]

        bar_chart = BarChart()
        bar_chart.title = "Income vs Expense"
        bar_chart.y_axis.title = "Amount"

        data = Reference(summary_sheet, min_col=2, min_row=2, max_row=3)
        categories = Reference(summary_sheet, min_col=1, min_row=2, max_row=3)

        bar_chart.add_data(data, titles_from_data=False)
        bar_chart.set_categories(categories)

        summary_sheet.add_chart(bar_chart, "E2")

        workbook.save(file_path)

        return file_path
        